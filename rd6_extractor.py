"""
rd6_extractor.py
Extracts project data from Malath/Tawuniya policy PDFs, the RD6 Master Excel,
and insulation certificate PDFs.
"""
import re
import io
from datetime import datetime
import pdfplumber
import openpyxl


# ── Helpers ────────────────────────────────────────────────────────────────────

def _clean(text):
    return re.sub(r'\s+', ' ', text or '').strip()

def _fmt_date(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        if val.year < 1901:
            return ''
        return '{}/{}/{}'.format(val.day, val.month, val.year)
    s = str(val).strip().replace('.0', '')
    return '' if s in ('None', '') else s

def _is_tawuniya(text):
    return 'tawuniya' in text.lower()


def _fix_rtl(text):
    """
    Fix Arabic text reversed by pdfplumber's RTL handling.
    pdfplumber reverses both word order and character order for Arabic.
    Reversing both back gives correct readable Arabic.
    Only applied when text is predominantly Arabic.
    """
    if not text or not text.strip():
        return text
    arabic_chars = sum(1 for c in text if '\u0600' <= c <= '\u06ff')
    if arabic_chars / max(len(text), 1) < 0.35:
        return text  # mostly Latin — don't reverse
    words = text.strip().split()
    return ' '.join(''.join(reversed(w)) for w in reversed(words))


# ── Malath extraction ──────────────────────────────────────────────────────────

def _extract_malath(text):
    data = {'ins_type': 'Malath'}

    # IDI / reference number
    for pat in [r'Reference\s+Number\s*[:\-]?\s*(\d{5,7})',
                r'Reference Number\s+(\d{5,7})']:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data['idi_no'] = m.group(1).strip()
            break

    # Owner
    m = re.search(r'Premises Owner\s*\n([^\n]+)', text, re.IGNORECASE)
    if m:
        c = _clean(m.group(1))
        if len(c) > 4:
            data['owner'] = c

    # Project title
    m = re.search(r'Name of Project\s*\n([^\n]+)', text, re.IGNORECASE)
    if m:
        data['project_title'] = _clean(m.group(1))

    # Address
    m = re.search(r'Premises Location.*?Street\s*\n([^\n]+)', text,
                  re.IGNORECASE | re.DOTALL)
    if m:
        c = _clean(m.group(1))
        if len(c) > 4:
            data['address'] = c

    # Sum insured
    m = re.search(r'Estimated Full Rebuilding Cost of the Premises\s+([\d,]+)\s*SR',
                  text, re.IGNORECASE)
    if m:
        data['sum_insured'] = _clean(m.group(1))

    # Building type
    m = re.search(r'Building\s+Type\s+(residential|commercial|mixed)', text, re.IGNORECASE)
    data['building_type'] = m.group(1).capitalize() if m else 'Residential'

    # OCC date
    m = re.search(r'Estimated Date of Issuing the Occupancy Certificate\s*([\d/\-]+)',
                  text, re.IGNORECASE)
    if m:
        data['occ_date'] = _clean(m.group(1))

    return data


# ── Tawuniya extraction ────────────────────────────────────────────────────────
#
# pdfplumber reads the bilingual two-column table in visual left-to-right order,
# producing lines like:
#   "Premises Owner [ARABIC DATA] [ARABIC LABEL reversed]"
#   "Name of Project [ARABIC DATA] [ARABIC LABEL reversed]"
# Strategy: use English label to locate each line, capture text between
# the English label and the reversed Arabic label at end of line.

def _extract_tawuniya(text):
    data = {'ins_type': 'Tawuniya', 'idi_no': ''}

    # Owner: line contains "Premises Owner [arabic data] [arabic label]"
    # Arabic label at end = reversed "مالك المباني" → "ينابملا كلام"
    m = re.search(r'Premises Owner\s+([^\n]+?)\s+(?:ينابملا|National Address)', text)
    if m:
        c = _fix_rtl(_clean(m.group(1)))
        if len(c) > 2:
            data['owner'] = c

    # Project title: "Name of Project [arabic data] [arabic label]"
    # Arabic label at end = reversed "اسم المشروع" → "عورشملا مسا"
    m = re.search(r'Name of Project\s+([^\n]+?)\s+(?:عورشملا|Building Type)', text)
    if m:
        c = _fix_rtl(_clean(m.group(1)))
        if len(c) > 2:
            data['project_title'] = c

    # Address/City: appears on its own line after "Premises Location:"
    # Line structure: "Premises Location: :ينابملا عقوم\n[CITY]\nCity/ District/ Street"
    m = re.search(r'Premises Location:[^\n]*\n([^\n]+)\n(?:City)', text)
    if m:
        c = _fix_rtl(_clean(m.group(1)))
        if len(c) > 1:
            data['address'] = c
    else:
        # Fallback: City Name line
        m = re.search(r'City Name\s+([^\n]+?)\s+(?:ةنيدملا|Zip)', text)
        if m:
            data['address'] = _fix_rtl(_clean(m.group(1)))

    # Sum insured (numeric — extracts cleanly)
    m = re.search(r'Estimated Full Rebuilding[^\d]*([\d,]+(?:\.\d+)?)',
                  text, re.IGNORECASE)
    if m:
        raw = m.group(1).replace(',', '')
        try:    data['sum_insured'] = '{:,}'.format(int(float(raw)))
        except: data['sum_insured'] = m.group(1)

    # OCC date — appears on line before 'Occupancy Certificate' label:
    # 'Estimated Date of Issuing of the 2025-10-09 00:00:00.0 [arabic]'
    # 'Occupancy Certificate'
    m = re.search(r'Estimated Date of Issuing of the\s+(\d{4}-\d{2}-\d{2})', text, re.IGNORECASE)
    if m:
        parts = m.group(1).split('-')
        if len(parts) == 3:
            data['occ_date'] = '{}/{}/{}'.format(int(parts[2]), int(parts[1]), parts[0])

    data['building_type'] = 'Residential'
    data['nt_ft']   = ''   # No NT/FT for Tawuniya
    data['taw_pol'] = ''   # NOT in PDF — engineer enters manually (it's the filename)
    return data


# ── Main PDF extraction entry point ───────────────────────────────────────────

def extract_from_policy_pdf(pdf_path):
    """Auto-detect Malath vs Tawuniya and extract accordingly."""
    full_text = ''
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            full_text += (page.extract_text() or '') + '\n'
    if _is_tawuniya(full_text):
        return _extract_tawuniya(full_text)
    else:
        return _extract_malath(full_text)


# ── Insulation certificate date extraction ─────────────────────────────────────

def extract_date_from_cert(cert_bytes, cert_filename):
    """
    Try to extract the document issue date from an insulation certificate PDF.
    Looks for the Gregorian date in Arabic format: م18/08/2025 (م immediately before date).
    Returns date as string d/m/yyyy, or '' if extraction fails.
    Only attempts extraction on PDF files.
    """
    import pathlib
    ext = pathlib.Path(cert_filename).suffix.lower()
    if ext != '.pdf':
        return ''
    try:
        full_text = ''
        with pdfplumber.open(io.BytesIO(cert_bytes)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or '') + '\n'
        # Match م immediately followed by date (no space) — avoids م 4:49 timestamp
        m = re.search(u'\u0645' + r'(\d{1,2}/\d{2}/\d{4})', full_text)
        if m:
            parts = m.group(1).split('/')
            return '{}/{}/{}'.format(int(parts[0]), int(parts[1]), parts[2])
        return ''
    except Exception:
        return ''


# ── Excel lookup ───────────────────────────────────────────────────────────────

def lookup_from_excel(excel_path, idi_no):
    """Lookup all fields from RD6 Master Excel by IDI_No."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']
    headers = [cell.value for cell in ws[1]]

    def ci(name):
        try: return headers.index(name)
        except ValueError: return None

    target = str(idi_no).strip().replace('.0', '')
    found_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = str(row[0]).strip().replace('.0', '') if row[0] is not None else ''
        if v == target:
            found_row = row
            break
    if found_row is None:
        return {}

    def get(name):
        c = ci(name)
        if c is None or c >= len(found_row): return ''
        v = found_row[c]
        if isinstance(v, datetime): return _fmt_date(v)
        return str(v).strip().replace('.0', '') if v is not None else ''

    def get_raw(name):
        c = ci(name)
        return found_row[c] if c is not None and c < len(found_row) else None

    visits = []
    for ordinal in ['1st', '2nd', '3rd', '4th', '5th', '6th', '7th']:
        ref  = get('{0}Visit_Ref'.format(ordinal))
        date = _fmt_date(get_raw('{0}Visit_date'.format(ordinal)))
        isp  = get('{0}Visit_isp'.format(ordinal))
        part = get('{0}Visit_part'.format(ordinal) if ordinal != '7th' else '7thVisit_part2')
        if ref and ref not in ('None', ''):
            visits.append({'ref': ref, 'date': date, 'inspector': isp, 'part': part})

    taw_pol = ''
    if 'Tuw-Mlth' in wb.sheetnames:
        ws2 = wb['Tuw-Mlth']
        for row in ws2.iter_rows(min_row=2, values_only=True):
            p = str(row[1]).strip().replace('.0', '') if row[1] is not None else ''
            if p == target:
                taw_pol = str(row[2]).strip().replace('.0', '') if row[2] else ''
                break
    if not taw_pol:
        taw_pol = get('Taw Pol.')

    return {
        'nt_ft':             get('NT/FT') or 'NT',
        'eng_full':          get('Eng'),
        'project_title':     get('ProjectTitle'),
        'address':           get('Address'),
        'owner':             get('Owner'),
        'start_date':        _fmt_date(get_raw('StartDate')),
        'finish_date':       _fmt_date(get_raw('FinishDate')),
        'last_visit_date':   _fmt_date(get_raw('Last_VisitDate')),
        'occ_date':          _fmt_date(get_raw('OCCDate')),
        'sum_insured':       get('TotCostRD0'),
        'rd0_ref':           get('RD0_Ref'),
        'rd0_date':          _fmt_date(get_raw('RD0Date')),
        'taw_pol':           taw_pol,
        'visits':            visits,
        'reservations_note': get('ReservationsNote'),
        'missing_doc':       get('MissingDoc'),
    }


# ── Reference builder ──────────────────────────────────────────────────────────

def build_rd6_reference(eng_full, nt_ft, idi_no, taw_pol, ins_type='Malath'):
    """
    Malath:   {INITIALS}-RD6-{NT/FT}{IDI_NO}-{TAW_POL}-1
    Tawuniya: {INITIALS}-RD6-{TAW_POL_NO}-01
    """
    parts = eng_full.strip().split()
    if len(parts) >= 2:
        initials = (parts[0][0] + parts[1][:2]).upper()
    elif parts:
        initials = parts[0][:3].upper()
    else:
        initials = 'ENG'

    idi = str(idi_no).strip().replace('.0', '')
    taw = str(taw_pol).strip().replace('.0', '')
    nt  = str(nt_ft).strip()

    if ins_type == 'Tawuniya':
        # Reference uses Tawuniya policy no, no NT/FT prefix
        return '{}-RD6-{}-01'.format(initials, taw) if taw else '{}-RD6-{}-01'.format(initials, idi)
    else:
        if taw:
            return '{}-RD6-{}{}-{}-1'.format(initials, nt, idi, taw)
        else:
            return '{}-RD6-{}{}-1'.format(initials, nt, idi)


# ── Engineer team lookup from IDI_Team.xlsx ────────────────────────────────────

def load_engineer_team(excel_path):
    """
    Load all engineers from IDI_Team.xlsx (Current Staff sheet).
    Returns dict: {full_name: {email, phone, degree, phase}}
    Phone numbers stored as integers in Excel — format as +966XXXXXXXXX.
    """
    import re
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['ENGs']
    except Exception:
        return {}

    team = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        name  = str(row[1]).strip()
        email = str(row[2]).strip() if row[2] else ''
        phone_raw = row[4]

        # Format phone: integer like 546380314 → +966546380314
        if phone_raw and str(phone_raw).strip() not in ('', '\xa0', 'None'):
            p = str(phone_raw).strip().replace('\xa0', '')
            # Extract first 9-digit number from the field
            digits = re.search(r'(5\d{8})', p.replace(' ','').replace('+966',''))
            if digits:
                phone = '+966' + digits.group(1)
            elif p.isdigit() and len(p) == 9:
                phone = '+966' + p
            else:
                phone = p  # keep as-is if format unclear
        else:
            phone = ''

        if name:
            team[name] = {
                'email': email,
                'phone': phone,
                'degree': 'Civil Engineering Bachelor',
                'phase': 'Senior',
            }
    return team
